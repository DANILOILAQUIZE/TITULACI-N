from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.generic import ListView
from django.contrib.auth.hashers import make_password
from .models import Grado, Periodo, Paralelo, PadronElectoral, CredencialUsuario
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import io
import random
import string
from django.core.mail import send_mail

#CRUD GRADOS
def generar_credenciales(request):
    # Inicializar el diccionario de credenciales generadas
    if not hasattr(request, 'session'):
        request.session = {}
    if 'credenciales_generadas' not in request.session:
        request.session['credenciales_generadas'] = {}
    
    if request.method == 'POST':
        # Obtener todos los usuarios del padrón que no tienen credencial
        usuarios_sin_credencial = PadronElectoral.objects.filter(
            credencial__isnull=True
        )
        
        credenciales_generadas = []
        
        for padron in usuarios_sin_credencial:
            # Generar una contraseña aleatoria de 8 caracteres
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
            
            try:
                # Generar una contraseña aleatoria de 8 caracteres
                password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                
                # Crear la nueva credencial con la contraseña en texto plano
                credencial = CredencialUsuario(
                    padron=padron,
                    usuario=padron.cedula,
                    estado='activo',
                    _contrasena_plana=password  # Establecer la contraseña en texto plano
                )
                
                # Guardar el objeto para generar el hash
                credencial.save()  # Esto generará automáticamente el hash en contrasena_encriptada
                
                # Guardar la contraseña en la sesión
                request.session['credenciales_generadas'][str(credencial.id)] = password
                request.session.modified = True
                
                credenciales_generadas.append({
                    'id': credencial.id,
                    'nombre': f"{padron.nombre} {padron.apellidos}",
                    'cedula': padron.cedula,
                    'contrasena': password,
                    'correo': padron.correo,
                    'estado': 'activo'
                })
                
                print(f"[DEBUG] Credencial creada para {padron.cedula} con contraseña: {password}")
                
            except Exception as e:
                print(f"[ERROR] Error al crear credencial para {padron.cedula}: {e}")
        
        if credenciales_generadas:
            messages.success(request, f'Se generaron {len(credenciales_generadas)} nuevas credenciales.')
            return render(request, 'padron/credenciales.html', {
                'credenciales': credenciales_generadas,
                'mostrar_credenciales': True
            })
        else:
            messages.info(request, 'No hay usuarios sin credenciales para generar.')
    
    # Si es GET o no se generaron credenciales, mostrar todas las credenciales
    credenciales = []
    for cred in CredencialUsuario.objects.select_related('padron').all():
        # Obtener la contraseña de la sesión si existe
        contrasena_plana = request.session.get('credenciales_generadas', {}).get(str(cred.id), "")
        
        print(f"[DEBUG] Mostrando credencial para {cred.padron.cedula}. Contraseña: {contrasena_plana}")
            
        credenciales.append({
            'id': cred.id,
            'nombre': f"{cred.padron.nombre} {cred.padron.apellidos}",
            'cedula': cred.padron.cedula,
            'contrasena': contrasena_plana,
            'correo': cred.padron.correo,
            'estado': cred.estado
        })
            
        return render(request, 'padron/credenciales.html', {
            'credenciales': credenciales,
            'mostrar_envio': True
        })

    # Si es GET, mostrar la lista de credenciales existentes
    credenciales = []
    for cred in CredencialUsuario.objects.select_related('padron').all():
        # Asegurarse de que tenemos la contraseña en texto plano
        contrasena_plana = cred.get_contrasena_plana  # Sin paréntesis, es una propiedad
        credenciales.append({
            'id': cred.id,
            'nombre': f"{cred.padron.nombre} {cred.padron.apellidos}",
            'cedula': cred.padron.cedula,
            'contrasena': contrasena_plana,  # Usar la contraseña en texto plano
            'correo': cred.padron.correo,
            'estado': cred.estado
        })
        
    return render(request, 'padron/credenciales.html', {
        'credenciales': credenciales,
        'mostrar_envio': False
    })

def enviar_credenciales(request):
    if request.method == 'POST':
        print('=== INICIO DE ENVÍO DE CREDENCIALES ===')
        print('POST data:', request.POST)
        print('IDs seleccionados:', request.POST.getlist('credenciales'))
        
        try:
            # Verificar que hay credenciales seleccionadas
            ids_seleccionados = request.POST.getlist('credenciales')
            if not ids_seleccionados:
                messages.error(request, 'Debe seleccionar al menos una credencial para enviar')
                print('=== FIN - NO SE SELECCIONARON CREDENCIALES ===')
                return redirect('generar_credenciales')
            
            try:
                # Obtener las credenciales seleccionadas
                credenciales = CredencialUsuario.objects.filter(id__in=ids_seleccionados).select_related('padron')
                print(f'Número de credenciales encontradas: {credenciales.count()}')
                
                if not credenciales.exists():
                    raise ValueError('No se encontraron credenciales válidas para enviar')
                    
                # Enviar correos electrónicos
                for credencial in credenciales:
                    print(f'=== PROCESANDO CREDENCIAL: {credencial.padron.cedula} ===')
                    print(f'Nombre: {credencial.padron.nombre} {credencial.padron.apellidos}')
                    print(f'Correo: {credencial.padron.correo}')
                    print(f'Estado: {credencial.estado}')
                    
                    try:
                        # Validar el formato del correo
                        from django.core.validators import validate_email
                        try:
                            validate_email(credencial.padron.correo)
                            print('Correo válido')
                        except Exception as e:
                            print(f'Error de validación de correo: {str(e)}')
                            raise ValueError(f'Correo electrónico inválido: {credencial.padron.correo}')
                        
                        # Obtener la contraseña de la sesión
                        contrasena_plana = request.session.get('credenciales_generadas', {}).get(str(credencial.id), "")
                        
                        if not contrasena_plana:
                            print(f"[ADVERTENCIA] No se encontró contraseña en texto plano para la credencial {credencial.id}")
                            contrasena_plana = "[CONTRASEÑA NO DISPONIBLE]"
                        
                        subject = 'Credenciales de acceso al Sistema de Votación de la Unidad Educativa Riobamba'
                        message = f"""Hola {credencial.padron.nombre} {credencial.padron.apellidos},

Tus credenciales de acceso al sistema son las siguientes:

Usuario (Cédula): {credencial.padron.cedula}
Contraseña: {contrasena_plana}

Por seguridad, te recomendamos cambiar tu contraseña después de iniciar sesión por primera vez.

Atentamente,
Consejo Electoral - Unidad Educativa Riobamba"""
                        
                        print('Enviando correo...')
                        send_mail(
                            subject,
                            message,
                            'riobamba@aplicacionesutc.com',  # Remitente consistente
                            [credencial.padron.correo],
                            fail_silently=False,  # Cambiado a False para ver errores
                        )
                        print('Correo enviado exitosamente')
                        
                        # No cambiar el estado al enviar el correo
                        # El estado se actualizará cuando el usuario cambie su contraseña
                        print('Correo enviado exitosamente sin cambiar el estado')
                        
                    except Exception as e:
                        import traceback
                        print('=== ERROR DETALLADO ===')
                        print(f'Tipo de error: {type(e).__name__}')
                        print(f'Mensaje de error: {str(e)}')
                        print('Traceback:')
                        print(traceback.format_exc())
                        print('=== FIN DEL ERROR ===')
                        
                        messages.warning(request, f'Error al enviar correo a {credencial.padron.correo}: {str(e)}')
                        print(f'Error al enviar correo a {credencial.padron.correo}: {str(e)}')
                        
                # Si no hubo errores críticos, mostrar mensaje de éxito general
                messages.success(request, 'Credenciales enviadas exitosamente. Algunos correos pueden haber fallado.')
                print('=== FIN - PROCESO COMPLETADO ===')
                return redirect('generar_credenciales')
                
            except ValueError as ve:
                print(f'Error de validación: {str(ve)}')
                messages.error(request, str(ve))
                print('=== FIN - ERROR DE VALIDACIÓN ===')
                return redirect('generar_credenciales')
                
        except Exception as e:
            print(f'Error inesperado: {str(e)}')
            import traceback
            print('=== ERROR INESPERADO ===')
            print(f'Tipo de error: {type(e).__name__}')
            print(f'Mensaje de error: {str(e)}')
            print('Traceback:')
            print(traceback.format_exc())
            print('=== FIN DEL ERROR ===')
            messages.error(request, 'Ocurrió un error inesperado al enviar las credenciales')
            print('=== FIN - ERROR INESPERADO ===')
            return redirect('generar_credenciales')
    
    return redirect('generar_credenciales')

#CRUD GRADOS
class GradoListView(ListView):
    model = Grado
    template_name = 'grados/agregarGrado.html'
    context_object_name = 'grados'
    paginate_by = 10

    def get_queryset(self):
        return Grado.objects.select_related('periodo').prefetch_related('paralelos').order_by('nombre')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # SOLUCIÓN: Usar el último período por fecha de inicio
        context['periodo_actual'] = Periodo.objects.order_by('-fecha_inicio').first()
        return context

def agregar_grado(request):
    # SOLUCIÓN: Obtener el período más reciente por fecha de inicio
    periodo_actual = Periodo.objects.order_by('-fecha_inicio').first()
    
    if not periodo_actual:
        messages.error(request, 'No hay períodos configurados en el sistema')
        return redirect('listar_grados')
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        
        if not nombre:
            messages.error(request, 'El nombre del grado es obligatorio')
            return render(request, 'grados/agregarGrado.html', {
                'periodo_actual': periodo_actual,
                'nombre': nombre
            })
        
        try:
            Grado.objects.create(
                nombre=nombre,
                periodo=periodo_actual
            )
            messages.success(request, 'Grado creado exitosamente!')
            return redirect('listar_grados')
        except Exception as e:
            messages.error(request, f'Error al crear grado: {str(e)}')
            return render(request, 'grados/agregar_grado.html', {
                'periodo_actual': periodo_actual,
                'nombre': nombre
            })
    
    return render(request, 'grados/agregar_grado.html', {
        'periodo_actual': periodo_actual
    })


def editar_grado(request, id):
    grado = get_object_or_404(Grado, pk=id)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        
        if not nombre:
            messages.error(request, 'El nombre del grado es obligatorio')
            return render(request, 'grados/editar_grado.html', {
                'grado': grado
            })
        
        try:
            grado.nombre = nombre
            # No modificamos el periodo, se mantiene el que ya tenía
            grado.save()
            messages.success(request, 'Grado actualizado correctamente!')
            return redirect('listar_grados')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
            return render(request, 'grados/editar_grado.html', {
                'grado': grado
            })
    
    return render(request, 'grados/editar_grado.html', {
        'grado': grado
    })

def eliminar_grado(request, id):
    grado = get_object_or_404(Grado, pk=id)
    
    try:
        if grado.padronelectoral_set.exists():
            messages.error(request, 'No se puede eliminar: Existen estudiantes asociados')
        elif grado.paralelos.exists():
            grado.paralelos.all().delete()
            grado.delete()
            messages.success(request, 'Grado y sus paralelos eliminados correctamente')
        else:
            grado.delete()
            messages.success(request, 'Grado eliminado correctamente')
    except Exception as e:
        messages.error(request, f'Error al eliminar: {str(e)}')
    
    return redirect('listar_grados')

#CRUD PARALELOS
class ParaleloListView(ListView):
    model = Paralelo
    template_name = 'paralelos/agregarParalelo.html'
    context_object_name = 'paralelos'
    paginate_by = 15

    def get_queryset(self):
        return Paralelo.objects.select_related('grado', 'grado__periodo').order_by('grado__nombre', 'nombre')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['grados'] = Grado.objects.all()
        return context

def agregar_paralelo(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre').upper()
        grado_id = request.POST.get('grado')
        
        if not nombre or not grado_id:
            messages.error(request, 'Todos los campos son obligatorios')
            return redirect('listar_paralelos')
        
        try:
            Paralelo.objects.create(
                nombre=nombre,
                grado_id=grado_id
            )
            messages.success(request, 'Paralelo creado exitosamente!')
        except IntegrityError:
            messages.error(request, 'Este paralelo ya existe para el grado seleccionado')
        except Exception as e:
            messages.error(request, f'Error al crear paralelo: {str(e)}')
        
        return redirect('listar_paralelos')
    
    return render(request, 'paralelos/agregarParalelo.html', {
        'grados': Grado.objects.all()
    })

def editar_paralelo(request, id):
    paralelo = get_object_or_404(Paralelo, pk=id)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre').upper()
        grado_id = request.POST.get('grado')
        
        if not nombre or not grado_id:
            messages.error(request, 'Todos los campos son obligatorios')
            return redirect('listar_paralelos')
        
        try:
            paralelo.nombre = nombre
            paralelo.grado_id = grado_id
            paralelo.save()
            messages.success(request, 'Paralelo actualizado correctamente!')
        except IntegrityError:
            messages.error(request, 'Este paralelo ya existe para el grado seleccionado')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
        
        return redirect('listar_paralelos')
    
    return render(request, 'paralelo/modals/form_paralelo.html', {
        'paralelo': paralelo,
        'grados': Grado.objects.all()
    })

def eliminar_paralelo(request, id):
    paralelo = get_object_or_404(Paralelo, pk=id)
    
    if paralelo.padronelectoral_set.exists():
        messages.error(request, '¡No se puede eliminar! Existen estudiantes asociados')
    else:
        try:
            paralelo.delete()
            messages.success(request, 'Paralelo eliminado correctamente')
        except Exception as e:
            messages.error(request, f'Error al eliminar: {str(e)}')
    
    return redirect('listar_paralelos')



# CRUD PADRON ELECTORAL
ESTADOS = [
    ('activo', 'Activo'),
    ('inactivo', 'Inactivo'),
]

def gestion_padron(request):
    padron = PadronElectoral.objects.select_related('grado', 'paralelo', 'periodo').all().order_by('apellidos', 'nombre')
    grados = Grado.objects.all().prefetch_related('paralelos')
    periodos = Periodo.objects.all()
    paralelos = Paralelo.objects.all()
    
    # Obtener el período actual (el más reciente por fecha de inicio)
    periodo_actual = Periodo.objects.order_by('-fecha_inicio').first()
    
    context = {
        'padron': padron,
        'grados': grados,
        'periodos': periodos,
        'paralelos': paralelos,
        'ESTADOS': ESTADOS,
        'periodo_actual': periodo_actual,  # Añadimos el período actual al contexto
    }
    return render(request, 'padron/agregarPadron.html', context)

def agregar_estudiante(request):
    if request.method == 'POST':
        # Obtener datos del formulario
        cedula = request.POST.get('cedula')
        nombre = request.POST.get('nombre').upper()  # Convertir a mayúsculas como en paralelo
        apellidos = request.POST.get('apellidos').upper()
        correo = request.POST.get('correo')
        telefono = request.POST.get('telefono')
        grado_id = request.POST.get('grado')
        paralelo_id = request.POST.get('paralelo')
        periodo_id = request.POST.get('periodo_id')
        estado = request.POST.get('estado')
        
        # Validación básica
        if not all([cedula, nombre, apellidos, correo, grado_id, paralelo_id, estado]):
            messages.error(request, 'Todos los campos obligatorios deben ser completados')
            return redirect('gestion_padron')
        
        try:
            # Verificar si el estudiante ya existe
            if PadronElectoral.objects.filter(cedula=cedula).exists():
                raise IntegrityError('Ya existe un estudiante con esta cédula')
            
            # Si no se proporcionó periodo, usar el actual
            if not periodo_id:
                periodo = Periodo.objects.order_by('-fecha_inicio').first()
                if not periodo:
                    raise ValueError('No hay períodos definidos en el sistema')
                periodo_id = periodo.id
            
            # Crear el estudiante
            PadronElectoral.objects.create(
                cedula=cedula,
                nombre=nombre,
                apellidos=apellidos,
                correo=correo,
                telefono=telefono,
                grado_id=grado_id,
                paralelo_id=paralelo_id,
                periodo_id=periodo_id,
                estado=estado
            )
            
            messages.success(request, 'Estudiante agregado exitosamente!')
            return redirect('gestion_padron')
            
        except IntegrityError as e:
            messages.error(request, f'Error de integridad: {str(e)}')
        except Grado.DoesNotExist:
            messages.error(request, 'El grado seleccionado no existe')
        except Paralelo.DoesNotExist:
            messages.error(request, 'El paralelo seleccionado no existe')
        except Exception as e:
            messages.error(request, f'Error al agregar estudiante: {str(e)}')
        
        return redirect('gestion_padron')
    
    # Si no es POST, mostrar el formulario con los datos necesarios
    return render(request, 'tu_template.html', {
        'grados': Grado.objects.all(),
        'paralelos': Paralelo.objects.all(),
        'periodo_actual': Periodo.objects.order_by('-fecha_inicio').first(),
        'ESTADOS': PadronElectoral.ESTADOS  # Asumiendo que ESTADOS está definido en el modelo
    })

def editar_estudiante(request, estudiante_id):
    estudiante = get_object_or_404(PadronElectoral, id=estudiante_id)
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            cedula = request.POST.get('cedula')
            nombre = request.POST.get('nombre').upper()  # Consistente con agregar_estudiante
            apellidos = request.POST.get('apellidos').upper()
            correo = request.POST.get('correo')
            telefono = request.POST.get('telefono')
            estado = request.POST.get('estado')  # Campo crítico
            grado_id = request.POST.get('grado')
            paralelo_id = request.POST.get('paralelo')
            periodo_id = request.POST.get('periodo_id')  # Ajustado para consistencia

            # Validación básica
            if not all([cedula, nombre, apellidos, correo, grado_id, paralelo_id, estado]):
                messages.error(request, 'Todos los campos obligatorios deben ser completados')
                return redirect('gestion_padron')

            # Validar el campo estado
            if hasattr(PadronElectoral, 'ESTADOS'):  # Verificar si el modelo tiene ESTADOS definidos
                estados_validos = [estado_val[0] for estado_val in PadronElectoral.ESTADOS]
                if estado not in estados_validos:
                    messages.error(request, f'El estado "{estado}" no es válido')
                    return redirect('gestion_padron')
            else:
                # Si no hay ESTADOS definidos, asumir valores por defecto (ajusta según tu modelo)
                estados_validos = ['activo', 'inactivo']
                if estado not in estados_validos:
                    messages.error(request, f'El estado "{estado}" no es válido')
                    return redirect('gestion_padron')

            # Asignar valores al estudiante
            estudiante.cedula = cedula
            estudiante.nombre = nombre
            estudiante.apellidos = apellidos
            estudiante.correo = correo
            estudiante.telefono = telefono
            estudiante.estado = estado  # Asignar el estado validado
            estudiante.grado = Grado.objects.get(id=grado_id)
            estudiante.paralelo = Paralelo.objects.get(id=paralelo_id)

            # Manejo del período (consistente con agregar_estudiante)
            if not periodo_id:
                periodo = Periodo.objects.order_by('-fecha_inicio').first()
                if not periodo:
                    messages.error(request, 'No hay períodos definidos en el sistema')
                    return redirect('gestion_padron')
                estudiante.periodo = periodo
            else:
                estudiante.periodo = Periodo.objects.get(id=periodo_id)

            # Guardar cambios
            estudiante.save()
            messages.success(request, 'Estudiante actualizado exitosamente!')
        
        except Grado.DoesNotExist:
            messages.error(request, 'El grado seleccionado no existe')
        except Paralelo.DoesNotExist:
            messages.error(request, 'El paralelo seleccionado no existe')
        except Periodo.DoesNotExist:
            messages.error(request, 'El período seleccionado no existe')
        except Exception as e:
            messages.error(request, f'Error al actualizar estudiante: {str(e)}')
        
        return redirect('gestion_padron')

    return render(request, 'editar_estudiante.html', {
        'estudiante': estudiante,
        'grados': Grado.objects.all(),
        'paralelos': Paralelo.objects.all(),
        'periodo_actual': Periodo.objects.order_by('-fecha_inicio').first(),
        'ESTADOS': getattr(PadronElectoral, 'ESTADOS', [('activo', 'Activo'), ('inactivo', 'Inactivo')])
    })
    

def eliminar_estudiante(request, estudiante_id):
    estudiante = get_object_or_404(PadronElectoral, id=estudiante_id)
    
    try:
        estudiante.delete()
        messages.success(request, 'Estudiante eliminado exitosamente!')
    except Exception as e:
        messages.error(request, f'Error al eliminar estudiante: {str(e)}')
    
    return redirect('gestion_padron')

def cargar_paralelos(request):
    grado_id = request.GET.get('grado_id')
    paralelos = Paralelo.objects.filter(grado_id=grado_id).order_by('nombre')
    
    data = {
        'paralelos': [{'id': p.id, 'nombre': p.nombre} for p in paralelos]
    }
    return JsonResponse(data)


# FORMATO PADORN ELECTORAL
def exportar_padron_excel(request):
    # Crear el libro de trabajo y la hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "PADRON_ELECTORAL_FORMATO"
    
    # Encabezados
    headers = [
        'Cédula', 'Apellidos', 'Nombres', 'Grado', 'Paralelo', 
        'Correo Electrónico', 'Teléfono(Opcional)'
    ]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws.column_dimensions[col_letter].width = 20
    
    # Datos
    estudiantes = PadronElectoral.objects.select_related('grado', 'paralelo', 'periodo').all()
    
    for row_num, estudiante in enumerate(estudiantes, 2):
        # Asegurarse de que las columnas coincidan con los encabezados
        ws[f'A{row_num}'] = estudiante.cedula
        ws[f'B{row_num}'] = estudiante.apellidos
        ws[f'C{row_num}'] = estudiante.nombre
        ws[f'D{row_num}'] = estudiante.grado.nombre if estudiante.grado else ''
        ws[f'E{row_num}'] = estudiante.paralelo.nombre if estudiante.paralelo else ''
        ws[f'F{row_num}'] = estudiante.correo if estudiante.correo else ''
        ws[f'G{row_num}'] = estudiante.telefono if estudiante.telefono else ''
    
    # Preparar la respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=PADRON_ELECTORAL_FORMATO.xlsx'
    
    # Guardar el libro en la respuesta
    with io.BytesIO() as buffer:
        wb.save(buffer)
        response.write(buffer.getvalue())
    
    return response

# CARGAR EL PADRON ELECTORAL DESDE UN ARCHIVO EXCEL
def importar_padron_excel(request):
    print("DEBUG: Iniciando importación de archivo Excel")
    if request.method != 'POST':
        print("DEBUG: Error - No es una petición POST")
        messages.error(request, 'Método no permitido')
        return redirect('gestion_padron')
        
    if not request.FILES.get('archivo_excel'):
        print("DEBUG: Error - No se recibió ningún archivo")
        messages.error(request, 'No se ha proporcionado ningún archivo')
        return redirect('gestion_padron')
    
    archivo = request.FILES['archivo_excel']
    print(f"DEBUG: Archivo recibido: {archivo.name}, Tamaño: {archivo.size} bytes")
    
    # Verificar que el archivo sea un Excel
    if not archivo.name.endswith(('.xlsx', '.xls')):
        print(f"DEBUG: Error - Formato de archivo no soportado: {archivo.name}")
        messages.error(request, 'El archivo debe ser de tipo Excel (.xlsx o .xls)')
        return redirect('gestion_padron')
    
    try:
        print("DEBUG: Intentando cargar el archivo Excel...")
        # Cargar el archivo Excel
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active
        print(f"DEBUG: Archivo cargado. Hojas: {wb.sheetnames}, Hoja activa: {ws.title}")
        
        # Verificar que el archivo no esté vacío
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"DEBUG: Filas: {max_row}, Columnas: {max_col}")
        
        if max_row <= 1:
            print("DEBUG: Error - El archivo está vacío o solo tiene encabezados")
            messages.error(request, 'El archivo está vacío o solo contiene encabezados')
            return redirect('gestion_padron')
            
        # Mostrar encabezados para depuración
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"DEBUG: Encabezados encontrados: {headers}")
        
        # Obtener el período activo actual
        periodo_actual = Periodo.objects.filter(estado='activo').order_by('-fecha_inicio').first()
        
        if not periodo_actual:
            messages.error(request, 'No hay un período académico activo. Por favor, cree un período activo primero.')
            return redirect('gestion_padron')
            
        print(f"DEBUG: Usando período académico: {periodo_actual.nombre}")
        print(f"DEBUG: Fecha de inicio: {periodo_actual.fecha_inicio}, Fecha de fin: {periodo_actual.fecha_fin}")
        
        # Contadores para estadísticas
        registros_procesados = 0
        registros_omitidos = 0
        
        with transaction.atomic():
            # Empezar desde la segunda fila (asumiendo que la primera es el encabezado)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Validar que la fila tenga suficientes columnas (mínimo 6 columnas requeridas)
                    if len(row) < 6:  # Reducido de 7 a 6 ya que el teléfono es opcional
                        registros_omitidos += 1
                        print(f"DEBUG: Fila {row_num} omitida - No tiene suficientes columnas")
                        continue
                        
                    # Obtener los valores de cada columna
                    # Orden: Cédula | Apellidos | Nombres | Grado | Paralelo | Correo | Teléfono (opcional)
                    try:
                        cedula = str(row[0]).strip() if row[0] else None
                        apellidos = str(row[1]).strip().upper() if row[1] else None
                        nombre = str(row[2]).strip().upper() if row[2] else None
                        grado_nombre = str(row[3]).strip().upper() if row[3] else None
                        paralelo_nombre = str(row[4]).strip().upper() if row[4] else None
                        correo = str(row[5]).strip().lower() if row[5] else None
                        telefono = str(row[6]).strip() if len(row) > 6 and row[6] else None
                        
                        print(f"DEBUG: Fila {row_num} - Datos leídos: {cedula}, {apellidos}, {nombre}, {grado_nombre}, {paralelo_nombre}, {correo}, {telefono}")
                    except Exception as e:
                        print(f"DEBUG: Error al procesar fila {row_num}: {str(e)}")
                        registros_omitidos += 1
                        continue
                    
                    # Validar campos obligatorios
                    campos_requeridos = [
                        ('Cédula', cedula),
                        ('Apellidos', apellidos),
                        ('Nombres', nombre),
                        ('Grado', grado_nombre),
                        ('Paralelo', paralelo_nombre),
                        ('Correo', correo)
                    ]
                    
                    campos_faltantes = [nombre for nombre, valor in campos_requeridos if not valor]
                    if campos_faltantes:
                        registros_omitidos += 1
                        print(f"DEBUG: Fila {row_num} omitida - Faltan campos obligatorios: {', '.join(campos_faltantes)}")
                        continue
                    
                    # Validar formato de cédula (10 dígitos)
                    if not (cedula.isdigit() and len(cedula) == 10):
                        print(f"DEBUG: Fila {row_num} omitida - Cédula inválida: {cedula}")
                        registros_omitidos += 1
                        continue
                    
                    # Validar formato de correo electrónico
                    if '@' not in correo or '.' not in correo.split('@')[-1]:
                        print(f"DEBUG: Fila {row_num} omitida - Correo inválido: {correo}")
                        registros_omitidos += 1
                        continue
                    
                    # Obtener o crear grado y paralelo
                    try:
                        # Limpiar y estandarizar nombres
                        grado_nombre = str(grado_nombre).strip().upper()
                        paralelo_nombre = str(paralelo_nombre).strip().upper()
                        
                        print(f"DEBUG: Procesando - Grado: {grado_nombre}, Paralelo: {paralelo_nombre}")
                        
                        # Obtener o crear el grado
                        grado, created = Grado.objects.get_or_create(
                            nombre=grado_nombre,
                            defaults={'periodo': periodo_actual}
                        )
                        
                        if created:
                            print(f"DEBUG: Nuevo grado creado: {grado.nombre}")
                        
                        # Obtener o crear el paralelo
                        paralelo, created = Paralelo.objects.get_or_create(
                            nombre=paralelo_nombre,
                            grado=grado
                        )
                        
                        if created:
                            print(f"DEBUG: Nuevo paralelo creado: {paralelo.nombre} para el grado {grado.nombre}")
                        
                        # Crear o actualizar estudiante
                        estudiante, created = PadronElectoral.objects.update_or_create(
                            cedula=cedula,
                            defaults={
                                'nombre': nombre,
                                'apellidos': apellidos,
                                'grado': grado,
                                'paralelo': paralelo,
                                'periodo': periodo_actual,
                                'correo': correo,
                                'telefono': telefono,
                                'estado': 'activo'
                            }
                        )
                        
                        if created:
                            print(f"DEBUG: Nuevo estudiante creado: {estudiante.apellidos} {estudiante.nombre}")
                        else:
                            print(f"DEBUG: Estudiante actualizado: {estudiante.apellidos} {estudiante.nombre}")
                        
                        registros_procesados += 1
                        
                    except Exception as e:
                        registros_omitidos += 1
                        continue
                        
                except Exception as e:
                    registros_omitidos += 1
                    continue
            
            # Mensaje de éxito con estadísticas
            if registros_procesados > 0:
                mensaje_exito = (
                    f'<strong>¡Importación completada con éxito!</strong><br>'
                    f'<strong>Período académico:</strong> {periodo_actual.nombre}<br>'
                    f'<strong>Registros procesados:</strong> {registros_procesados}<br>'
                    f'<strong>Registros omitidos:</strong> {registros_omitidos}'
                )
                messages.success(request, mensaje_exito, extra_tags='alert-success')
            else:
                mensaje_error = (
                    '<strong>No se pudo procesar ningún registro. Por favor verifique:</strong><br>'
                    '1. El formato del archivo debe ser: Cédula | Apellidos | Nombres | Grado | Paralelo | Correo | Teléfono (opcional)<br>'
                    '2. La primera fila debe contener los encabezados<br>'
                    '3. No debe haber filas vacías<br>'
                    '4. Las cédulas no deben estar duplicadas<br>'
                    '5. Los correos electrónicos deben tener un formato válido'
                )
                messages.error(request, mensaje_error, extra_tags='alert-danger')
                
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"DEBUG: Error al procesar el archivo: {str(e)}")
        print(f"DEBUG: Traceback completo: {error_trace}")
        
        messages.error(
            request, 
            f'Error al procesar el archivo: {str(e)}\n'
            'Asegúrese de que el archivo no esté abierto en otro programa, que el formato sea correcto '
            'y que tenga los permisos necesarios.'
        )
    
    print("DEBUG: Redirigiendo a la página de gestión de padrón")  # Debug
    return redirect('gestion_padron')
